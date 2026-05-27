import io
import os
import re
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Dict, List

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

load_dotenv()

EMAIL_COLUMNS = [
    "Call Recording Link",
    "Average Score",
    "Score Parameter Wise",
    "Strengths",
    "Improvement Areas",
    "Learnings",
]


def get_setting(name: str, default: str | None = None) -> str | None:
    value = os.getenv(name)
    if value:
        return value
    try:
        import streamlit as st
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return default


def _split_recipients(raw: str | None) -> List[str]:
    return [r.strip() for r in (raw or "").split(",") if r.strip()]


def _format_multiline(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    # Convert numbered paragraphs into separate lines: "1. ... 2. ..." -> "1. ...\n2. ..."
    text = re.sub(r"\s+(?=\d+[.)]\s+)", "\n", text)
    return text


def _normalize_average(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    ratio = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", text)
    if ratio:
        numerator = float(ratio.group(1))
        denominator = float(ratio.group(2))
        if denominator > 0:
            return f"{(numerator / denominator) * 10:.1f}"

    number = re.search(r"\d+(?:\.\d+)?", text)
    if number:
        return f"{float(number.group(0)):.1f}"

    return text


def build_excel(results: List[Dict]) -> bytes:
    """Build a polished XLSX report. CSV cannot store header color, bold text, or wrap settings."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Scores"

    header_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    header_font = Font(bold=True, color="000000")
    border_side = Side(style="thin", color="D9D9D9")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_idx, header in enumerate(EMAIL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(results, 2):
        for col_idx, header in enumerate(EMAIL_COLUMNS, 1):
            value = row.get(header, "")
            if header == "Average Score":
                value = _normalize_average(value)
            elif header not in {"Call Recording Link"}:
                value = _format_multiline(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if header == "Call Recording Link" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 34,  # Call Recording Link
        "B": 14,  # Average Score
        "C": 34,  # Score Parameter Wise
        "D": 45,  # Strengths
        "E": 55,  # Improvement Areas
        "F": 55,  # Learnings
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 120
    ws.row_dimensions[1].height = 28

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _batch_summary(results: List[Dict]) -> Dict[str, int]:
    return {
        "total": len(results),
        "full_analysis": sum(1 for r in results if r.get("Call Type") == "full_analysis"),
        "follow_up_only": sum(1 for r in results if r.get("Call Type") == "follow_up_only"),
        "not_worthy": sum(1 for r in results if r.get("Call Type") == "not_worthy"),
        "converted": sum(1 for r in results if r.get("Converted Status") == "Converted"),
    }


def _smtp_send(sender: str, password: str, recipients: List[str], message: MIMEMultipart) -> None:
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, password)
        server.sendmail(sender, recipients, message.as_string())


def send_report(results: List[Dict], batch_label: str = "Today") -> None:
    sender = get_setting("SENDER_EMAIL")
    password = get_setting("SENDER_PASSWORD")
    recipients = _split_recipients(get_setting("RECIPIENT_EMAILS", ""))

    if not sender:
        raise RuntimeError("SENDER_EMAIL is missing.")
    if not password:
        raise RuntimeError("SENDER_PASSWORD is missing. Use a Gmail App Password, not your normal Gmail password.")
    if not recipients:
        raise RuntimeError("RECIPIENT_EMAILS is missing.")

    excel_bytes = build_excel(results)
    summary = _batch_summary(results)

    body = f"""Hi,

Please find attached the EPFO call scoring report for {batch_label}.

Summary:
- Total calls processed: {summary['total']}
- Full analysis calls: {summary['full_analysis']}
- Follow-up only calls: {summary['follow_up_only']}
- Not worthy calls: {summary['not_worthy']}
- Converted calls: {summary['converted']}

Regards,
Rohit Sharma
"""

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"Call Scores Report - {batch_label}"
    msg.attach(MIMEText(body, "plain"))

    attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.set_payload(excel_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        f"attachment; filename=EduTap_Call_Scores_{batch_label.replace(' ', '_')}.xlsx",
    )
    msg.attach(attachment)

    _smtp_send(sender, password, recipients, msg)


def send_error_report(error_items: List[Dict], batch_label: str = "Today") -> None:
    """Send a simple error email for files that could not be processed."""
    if not error_items:
        return

    sender = get_setting("SENDER_EMAIL")
    password = get_setting("SENDER_PASSWORD")
    recipients = _split_recipients(get_setting("ERROR_RECIPIENT_EMAILS") or get_setting("RECIPIENT_EMAILS", ""))

    if not sender:
        raise RuntimeError("SENDER_EMAIL is missing.")
    if not password:
        raise RuntimeError("SENDER_PASSWORD is missing. Use a Gmail App Password, not your normal Gmail password.")
    if not recipients:
        raise RuntimeError("ERROR_RECIPIENT_EMAILS or RECIPIENT_EMAILS is missing.")

    lines = [
        "Hi,",
        "",
        f"Some EduTap call files could not be processed for {batch_label}.",
        "",
        "Failed items:",
    ]

    for index, item in enumerate(error_items, 1):
        lines.extend([
            "",
            f"{index}. File: {item.get('filename', 'Unknown file')}",
            f"Student number: {item.get('student_number', 'Unknown')}",
            f"Simple message: {item.get('simple_error', 'This file could not be processed.')}",
            f"Technical detail: {item.get('technical_error', 'Not available')}",
        ])

    lines.extend(["", "Regards,", "EduTap Call Scoring System"])
    body = "\n".join(lines)

    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"EduTap Call Processing Error - {batch_label}"
    msg.attach(MIMEText(body, "plain"))

    _smtp_send(sender, password, recipients, msg)
